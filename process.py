from openpyxl import load_workbook
import argparse
from config import get_config
from copy import copy

def xlWriteLine(sheet, content, irow):
    sheet.insert_rows(irow+1)
    for icol, source_cell in enumerate(content):
        # colchar = convertColumnLetter(icol)
        # sheet[f'{colchar}{irow+1}'] = icell.value
        target_cell = sheet.cell(irow+1, icol+1)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

parser = argparse.ArgumentParser('xlsx reader', add_help=False)
parser.add_argument('--cfg', default="default.yaml", type=str)
args = parser.parse_args()

cfg = get_config(args.cfg)

wtFile = load_workbook(cfg["Template"], data_only=True)
wtPointer = {}
sheetDict = {}
for isheet in cfg["Sheets"]:
    sheetDict[isheet["name"]] = {
        "prefix_lines": isheet["prefix_lines"],
        "posfix_string": isheet["posfix_string"],
    }
    wtPointer[isheet["name"]] = isheet["prefix_lines"]

print(f"Info: 共有{len(cfg['Files'])}个文件待处理")
for fileName in cfg["Files"]:
    iFile = load_workbook(fileName, data_only=True)
    fileSheetNames = iFile.sheetnames
    for sheetInfo in cfg["Sheets"]:
        sheetName = sheetInfo["name"]
        if sheetName not in fileSheetNames:
            print(f"Warning: {fileName} 文件中没有表格 {sheetName}")
        else:
            isheet = iFile[sheetName]
            wtSheet = wtFile[sheetName]
            rdpt = sheetDict[sheetName]["prefix_lines"]
            meetEnd=False
            for idx, irow in enumerate(isheet.rows):
                if idx < rdpt:
                    continue
                # 1 对应 string
                if isinstance(irow[0].value, str) and irow[0].value.replace(" ", "") == sheetDict[sheetName]["posfix_string"]:
                    meetEnd=True
                    break
                xlWriteLine(wtSheet, irow, wtPointer[sheetName])
                wtPointer[sheetName] += 1
            if sheetDict[sheetName]["posfix_string"] is not None and not meetEnd:
                print(f"Warning: 未在 {fileName} 的 {sheetName} 表中找到结束符 {sheetDict[sheetName]['posfix_string']}")
    iFile.close()

wtFile.save(cfg["Output"])
print(f"Info: 处理完成，已保存至{cfg['Output']}")
wtFile.close()