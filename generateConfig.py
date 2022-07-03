import argparse
from config import get_config
from openpyxl import load_workbook
import yaml
from collections import OrderedDict

parser = argparse.ArgumentParser('src-det', add_help=False)
parser.add_argument('tmp', type=str)
parser.add_argument('dst', type=str)
args = parser.parse_args()

cfg = {
    "Template": args.tmp,
    "Path": "~请填写需要统计的所有表格文件的表达式(eg: E:\目录\*.xlsx)",
    "Output": "~请填写统计后保存文件名(eg: E:\结果.xlsx)",
    "Sheets": [],
}

wtFile = load_workbook(args.tmp)
sheetNames = wtFile.sheetnames
for sheetName in sheetNames:
    ws = wtFile.get_sheet_by_name(sheetName)
    posfix = ws.cell(ws.max_row, 1).value
    if isinstance(posfix, str):
        posfix = posfix.replace(' ', '')
    cfg["Sheets"].append({
        "name": sheetName,
        "prefix_lines": "~请填写前缀说明行数，并确认终止符号是否正确",
        "posfix_string": posfix,
    })

with open(args.dst, "w", encoding='utf-8') as f:
    yaml.dump(cfg, f, allow_unicode=True, sort_keys=False)